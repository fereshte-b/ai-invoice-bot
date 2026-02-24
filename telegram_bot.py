import os
import json
import re
import base64
import logging
from datetime import datetime, timezone

from openai import OpenAI
from telegram import Update
from telegram.ext import ApplicationBuilder, MessageHandler, ContextTypes, filters

from openpyxl import Workbook, load_workbook

# -----------------------------
# Config & Logging
# -----------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
)
logger = logging.getLogger("ai-invoice-bot")

BOT_TOKEN = os.getenv("BOT_TOKEN")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")

PORT = int(os.getenv("PORT", "8080"))

# Excel path should be on Railway Volume mount (/data)
EXCEL_PATH = os.getenv("EXCEL_PATH", "/data/invoices.xlsx")

# Optional: Provide explicitly if you want webhook
# Example: https://ai-invoice-bot-production.up.railway.app
WEBHOOK_URL = os.getenv("WEBHOOK_URL")

# Some people store it like this on Railway:
# - RAILWAY_PUBLIC_DOMAIN (sometimes available) e.g. ai-invoice-bot-production.up.railway.app
# - RAILWAY_STATIC_URL (older examples)
RAILWAY_PUBLIC_DOMAIN = os.getenv("RAILWAY_PUBLIC_DOMAIN")
RAILWAY_STATIC_URL = os.getenv("RAILWAY_STATIC_URL")

if not BOT_TOKEN:
    raise ValueError("BOT_TOKEN not set")
if not OPENAI_API_KEY:
    raise ValueError("OPENAI_API_KEY not set")

client = OpenAI(api_key=OPENAI_API_KEY)

HEADERS = ["Date", "Supplier", "Net Total", "VAT"]


# -----------------------------
# Helpers
# -----------------------------
def ensure_excel_exists(path: str) -> None:
    """Create workbook with correct headers if file doesn't exist."""
    if os.path.exists(path):
        return
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"
    ws.append(HEADERS)
    wb.save(path)
    logger.info("Created new excel at %s", path)


def append_row_to_excel(path: str, row: dict) -> None:
    ensure_excel_exists(path)
    wb = load_workbook(path)
    ws = wb.active

    ws.append([
        row.get("date", ""),
        row.get("supplier", ""),
        row.get("net_total", ""),
        row.get("vat", ""),
    ])
    wb.save(path)


def extract_json_block(text: str) -> str:
    """
    Extract first JSON object from text.
    We force model to return JSON only, but this is extra safe.
    """
    text = text.strip()

    # If it's already pure JSON
    if text.startswith("{") and text.endswith("}"):
        return text

    # Try to find JSON object inside
    m = re.search(r"\{[\s\S]*\}", text)
    if not m:
        raise ValueError(f"No JSON found in model output: {text[:200]}")
    return m.group(0)


def normalize_number(x):
    """
    Convert strings like '13,050', '13.050', '13 050' to float if possible.
    Otherwise return original.
    """
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)

    s = str(x).strip()
    if not s:
        return None

    # Remove currency words/symbols (keep digits, comma, dot)
    s = re.sub(r"[^\d,.\-]", "", s)

    # Common cases:
    # - "13,050" could be 13050 (thousand separator) OR 13.05 (decimal). In invoices it's usually thousand.
    # We'll decide by last separator: if there are 3 digits after comma/dot and only one separator => thousand sep.
    # If there are 2 digits after => decimal.
    # If both comma and dot exist: assume one is thousand, one is decimal (common European).
    if "," in s and "." in s:
        # assume commas are thousand, dot decimal: "13,050.25" => 13050.25
        s = s.replace(",", "")
        try:
            return float(s)
        except:
            return x

    # Only comma present
    if "," in s and "." not in s:
        parts = s.split(",")
        if len(parts) == 2 and len(parts[1]) == 3:
            s2 = parts[0] + parts[1]  # thousand separator
            try:
                return float(s2)
            except:
                return x
        else:
            # treat as decimal comma
            s2 = s.replace(",", ".")
            try:
                return float(s2)
            except:
                return x

    # Only dot present
    if "." in s and "," not in s:
        parts = s.split(".")
        if len(parts) == 2 and len(parts[1]) == 3:
            s2 = parts[0] + parts[1]  # thousand separator
            try:
                return float(s2)
            except:
                return x
        try:
            return float(s)
        except:
            return x

    # Just digits
    try:
        return float(s)
    except:
        return x


def ai_extract_invoice(image_bytes: bytes) -> dict:
    """
    Use OpenAI Vision to extract invoice fields.
    Returns dict with:
      - invoice_date (string or null)
      - supplier (string or null)
      - net_total (number or string or null)
      - vat_amount (number or null)
    """
    b64 = base64.b64encode(image_bytes).decode("utf-8")

    prompt = """
You are extracting key fields from an invoice image.

Return ONLY valid JSON with EXACT keys:
{
  "invoice_date": "YYYY-MM-DD or null",
  "supplier": "string or null",
  "net_total": "number or string or null",
  "vat_amount": "number or null"
}

Rules:
- net_total should be the final amount to pay (grand total / total amount / net total depending on invoice language).
- If invoice_date cannot be found, use null.
- supplier should be vendor/seller/store name. If not found, null.
- vat_amount: if VAT is present with any numeric amount, return that numeric amount.
  If VAT is mentioned but amount is not visible, return 0.
  If VAT is NOT present at all, return null.
- Do NOT include any extra text. JSON ONLY.
"""

    resp = client.responses.create(
        model=os.getenv("OPENAI_MODEL", "gpt-4o-mini"),
        input=[
            {
                "role": "user",
                "content": [
                    {"type": "input_text", "text": prompt.strip()},
                    {"type": "input_image", "image_url": f"data:image/jpeg;base64,{b64}"},
                ],
            }
        ],
        temperature=0,
    )

    text = (resp.output_text or "").strip()
    json_str = extract_json_block(text)
    data = json.loads(json_str)

    # Basic normalization
    invoice_date = data.get("invoice_date")
    supplier = data.get("supplier")
    net_total = data.get("net_total")
    vat_amount = data.get("vat_amount")

    net_total_num = normalize_number(net_total)
    vat_amount_num = normalize_number(vat_amount)

    # VAT flag: if any VAT amount exists and > 0 => Yes, else No
    # If vat_amount is 0 (mentioned but not visible) => treat as Yes? user asked: "اگر مبلغ هست بزنه اگر نیست no"
    # So only "Yes" when numeric amount > 0.
    vat_flag = "Yes" if (vat_amount_num is not None and isinstance(vat_amount_num, float) and vat_amount_num > 0) else "No"

    return {
        "invoice_date": invoice_date,
        "supplier": (supplier or "").strip() if supplier else "",
        "net_total": net_total_num if isinstance(net_total_num, float) else (net_total or ""),
        "vat": vat_flag,
        "vat_amount": vat_amount_num,
    }


def to_yyyy_mm_dd(dt: datetime) -> str:
    return dt.strftime("%Y-%m-%d")


# -----------------------------
# Telegram Handlers
# -----------------------------
async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        if not update.message or not update.message.photo:
            return

        # Pick largest photo
        photo = update.message.photo[-1]
        file = await context.bot.get_file(photo.file_id)
        file_bytes = await file.download_as_bytearray()
        image_bytes = bytes(file_bytes)

        await update.message.reply_text("✅ عکس دریافت شد. دارم با AI تحلیل می‌کنم...")

        extracted = ai_extract_invoice(image_bytes)

        # Fallback date: message date (UTC) if invoice_date missing
        msg_dt = update.message.date
        if msg_dt is None:
            msg_dt = datetime.now(timezone.utc)

        date_str = extracted.get("invoice_date")
        if not date_str or str(date_str).lower() == "null":
            date_str = to_yyyy_mm_dd(msg_dt)

        row = {
            "date": date_str,
            "supplier": extracted.get("supplier", ""),
            "net_total": extracted.get("net_total", ""),
            "vat": extracted.get("vat", "No"),
        }

        append_row_to_excel(EXCEL_PATH, row)

        await update.message.reply_text("✅ اکسل بروزرسانی شد. فایل رو می‌فرستم...")

        # Send the updated excel back
        with open(EXCEL_PATH, "rb") as f:
            await update.message.reply_document(document=f, filename=os.path.basename(EXCEL_PATH))

    except Exception as e:
        logger.exception("Error in handle_photo")
        await update.message.reply_text(f"❌ خطا: {e}")


async def handle_other(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("فقط عکس فاکتور رو بفرست تا اکسل آپدیت بشه ✅")


# -----------------------------
# Webhook / Polling bootstrap
# -----------------------------
def resolve_webhook_url() -> str | None:
    if WEBHOOK_URL:
        return WEBHOOK_URL.rstrip("/")

    # Railway domain variants
    if RAILWAY_PUBLIC_DOMAIN:
        # might be just host without scheme
        host = RAILWAY_PUBLIC_DOMAIN.strip()
        if host.startswith("http://") or host.startswith("https://"):
            return host.rstrip("/")
        return f"https://{host}".rstrip("/")

    if RAILWAY_STATIC_URL:
        host = RAILWAY_STATIC_URL.strip()
        if host.startswith("http://") or host.startswith("https://"):
            return host.rstrip("/")
        return f"https://{host}".rstrip("/")

    return None


def main():
    ensure_excel_exists(EXCEL_PATH)

    app = ApplicationBuilder().token(BOT_TOKEN).build()

    app.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    app.add_handler(MessageHandler(~filters.PHOTO, handle_other))

    webhook_base = resolve_webhook_url()

    if webhook_base:
        # Use webhook (good for Railway)
        # Telegram requires HTTPS endpoint
        logger.info("Starting with webhook: %s", webhook_base)
        app.run_webhook(
            listen="0.0.0.0",
            port=PORT,
            url_path="webhook",
            webhook_url=f"{webhook_base}/webhook",
            allowed_updates=Update.ALL_TYPES,
        )
    else:
        # Fallback to polling (no domain needed)
        logger.warning("WEBHOOK_URL/RAILWAY_PUBLIC_DOMAIN not set -> starting polling.")
        app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
